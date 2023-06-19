import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os


resulting_file = 'Benchmarking/2023-06-16/Benchmarks.xlsx'
if os.path.exists(resulting_file):
	verify_run = input(f'Running this will overwite "{resulting_file}". Do you wish to continue? (y/n): ')
	if verify_run.lower() != 'y':
		print('Program Aborted')
		exit(1)

## Parse the text output of test and put it into an excel sheet
with pd.ExcelWriter(resulting_file, engine='openpyxl') as writer:
	with open('Benchmarking/2023-06-16/Test.txt', 'r') as f:
		# Read in as text
		file_string = f.read()
		tests = file_string.split('STARTING SBITS VARIABLE DATA TESTS.')
		for test in tests[1:]:
			# Parse out test info
			info = re.findall(r'VARDATA SIZE: (\d+)\n', test, re.DOTALL)[0]

			# Parse out stat table
			text_tables = re.findall(r'(Stats for 10000:.*?)(?=\n\nSTARTING SBITS|$)', test, re.DOTALL)
			text_tables = [x.strip() for x in text_tables]

			# Regex pattern to match each row of the table
			pattern = r"([\w| ]+):\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)|(Stats for \d+:)"

			for text_table in text_tables:
				# Extract the matches
				text_rows = re.findall(pattern, text_table)

				# Create a list to store the rows of the table
				table_rows = []

				# Iterate over the matches and create rows
				for text_row in text_rows:
					if text_row[-1] != '':
						table_rows.append([text_row[-1]] + ['' for num in text_row[1:-1]])
					else:
						table_rows.append([text_row[0]] + [int(num) for num in text_row[1:-1]])

				# Create the Pandas DataFrame
				columns = ["Stats", "Run1", "Run2", "Run3", "Avg"]
				df = pd.DataFrame(table_rows, columns=columns)

				# Create sheet name
				vardata_size = info
				sheet_name = f'newSD_sea100K_key=4_var={vardata_size}'

				# Write to excel workbook
				df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

wb = openpyxl.load_workbook(resulting_file)
sheets = wb.sheetnames

equations = [
		['Inserts', 'Queries', 'Writes', 'Buf Hits', 'Write Time', 'Insert Time', 'Read Time', 'Query Time', 'Num Reads', 'Inserts/sec', 'Queries/sec'],
		[10000, 10000, '=E3', '=E10', '=E7', '=K3/1000', '=E8', '=M3/1000', '=E9', '=G3/L3', '=H3/N3'],
		[20000, 20000, '=E13', '=E20', '=E17', '=K4/1000', '=E18', '=M4/1000', '=E19', '=G4/L4', '=H4/N4'],
		[30000, 30000, '=E23', '=E30', '=E27', '=K5/1000', '=E28', '=M5/1000', '=E29', '=G5/L5', '=H5/N5'],
		[40000, 40000, '=E33', '=E40', '=E37', '=K6/1000', '=E38', '=M6/1000', '=E39', '=G6/L6', '=H6/N6'],
		[50000, 50000, '=E43', '=E50', '=E47', '=K7/1000', '=E48', '=M7/1000', '=E49', '=G7/L7', '=H7/N7'],
		[60000, 60000, '=E53', '=E60', '=E57', '=K8/1000', '=E58', '=M8/1000', '=E59', '=G8/L8', '=H8/N8'],
		[70000, 70000, '=E63', '=E70', '=E67', '=K9/1000', '=E68', '=M9/1000', '=E69', '=G9/L9', '=H9/N9'],
		[80000, 80000, '=E73', '=E80', '=E77', '=K10/1000', '=E78', '=M10/1000', '=E79', '=G10/L10', '=H10/N10'],
		[90000, 90000, '=E83', '=E90', '=E87', '=K11/1000', '=E88', '=M11/1000', '=E89', '=G11/L11', '=H11/N11'],
		[100000, 100000, '=E93', '=E100', '=E97', '=K12/1000', '=E98', '=M12/1000', '=E99', '=G12/L12', '=H12/N12']
	]
ft = Font(bold=True)

## Put all the collected data into a neat table for each sheet
for sheet in sheets:
	ws = wb[sheet]
	for rn, row in enumerate(ws['G2':'Q12']):
		for cn, cell in enumerate(row):
			cell.value = equations[rn][cn]
			if rn == 0 or cn in [0,1]:
				cell.font = ft


## Put together data into tables for making charts
graph_sheet = wb.create_sheet('Charts', index=0)

# Inserts/sec as vardata size increases
var_sizes = ['0', '10', '50', '100', '500', '1000']
graph_sheet['O2'].value = 'IPS'
graph_sheet['O2'].font = ft
for offset, type in enumerate(var_sizes):
	cel = graph_sheet.cell(row=3+offset, column=15)
	cel.value = type
for col_offset, size in enumerate(var_sizes):
	graph_sheet.cell(row=3+col_offset, column=16).value = f"='newSD_sea100K_key=4_var={size}'!P12"

# Queries/sec as vardata size increases
graph_sheet['O23'].value = 'QPS'
graph_sheet['O23'].font = ft
for offset, type in enumerate(var_sizes):
	cel = graph_sheet.cell(row=24+offset, column=15)
	cel.value = type
for col_offset, size in enumerate(var_sizes):
	graph_sheet.cell(row=24+col_offset, column=16).value = f"='newSD_sea100K_key=4_var={size}'!Q12"


wb.save(resulting_file)
print('Done.')
